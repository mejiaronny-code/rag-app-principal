import io
import logging
import openpyxl
from typing import List, Tuple
import httpx
from bs4 import BeautifulSoup
import pypdf
from docx import Document as DocxDocument
from langchain.text_splitter import RecursiveCharacterTextSplitter
from services.ocr import extract_text_from_image_bytes, extract_text_from_scanned_pdf, is_pdf_scanned

logger = logging.getLogger(__name__)

# Chunking config
text_splitter = RecursiveCharacterTextSplitter(
    chunk_size=1000,
    chunk_overlap=200,
    length_function=len,
    separators=["\n\n", "\n", ". ", " ", ""],
)

# Magic bytes para validar tipos reales de archivo
MAGIC_BYTES = {
    b"%PDF": "pdf",
    b"PK\x03\x04": "docx",  # ZIP (docx, xlsx, etc.)
    b"\xff\xd8\xff": "jpg",
    b"\x89PNG": "png",
    b"GIF8": "gif",
    b"RIFF": "webp",
}


def detect_file_type(content: bytes) -> str:
    """Detecta el tipo real de archivo por sus magic bytes."""
    for magic, ftype in MAGIC_BYTES.items():
        if content[:len(magic)] == magic:
            return ftype
    # Si es texto plano
    try:
        content[:512].decode("utf-8")
        return "text"
    except UnicodeDecodeError:
        return "unknown"


def extract_text_from_pdf(content: bytes) -> str:
    """Extrae texto de PDF (con fallback a OCR si es escaneado)."""
    if is_pdf_scanned(content):
        logger.info("PDF detectado como escaneado, aplicando OCR...")
        return extract_text_from_scanned_pdf(content)
    
    reader = pypdf.PdfReader(io.BytesIO(content))
    texts = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text and text.strip():
            texts.append(f"[Página {i + 1}]\n{text.strip()}")
    return "\n\n".join(texts)


def extract_text_from_docx(content: bytes) -> str:
    """Extrae texto de un archivo DOCX."""
    doc = DocxDocument(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


async def extract_text_from_url(url: str) -> str:
    """Scrapea texto limpio de una URL usando BeautifulSoup + httpx."""
    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; RAGBot/1.0)",
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "es,en;q=0.9",
    }
    
    async with httpx.AsyncClient(
        timeout=15.0,
        follow_redirects=True,
        max_redirects=3,
        headers=headers,
    ) as client:
        response = await client.get(url)
        response.raise_for_status()
    
    # Solo procesar HTML/texto
    content_type = response.headers.get("content-type", "")
    if "text/html" not in content_type and "text/plain" not in content_type:
        raise ValueError(f"Tipo de contenido no soportado para scraping: {content_type}")
    
    soup = BeautifulSoup(response.text, "html.parser")
    
    # Eliminar scripts, estilos y nav
    for tag in soup(["script", "style", "nav", "footer", "header", "aside", "form"]):
        tag.decompose()
    
    # Extraer texto limpio
    text = soup.get_text(separator="\n", strip=True)
    
    # Limpiar líneas vacías múltiples
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return "\n".join(lines)


def chunk_text(text: str) -> List[str]:
    """Divide el texto en chunks con LangChain RecursiveCharacterTextSplitter."""
    if not text or not text.strip():
        raise ValueError("El texto extraído está vacío.")
    return text_splitter.split_text(text)


async def process_document(
    filename: str,
    content: bytes,
    file_type_hint: str,
) -> Tuple[str, List[str], str]:
    """
    Procesa un documento y retorna (texto_completo, chunks, tipo_detectado).
    """
    detected_type = detect_file_type(content)
    
    # Resolver tipo real
    if detected_type == "pdf" or file_type_hint == "pdf":
        doc_type = "pdf"
        text = extract_text_from_pdf(content)
    elif file_type_hint in ("xlsx", "xls"):
        doc_type = "xlsx"
        text = extract_text_from_excel(content)    
    elif detected_type == "docx" or file_type_hint in ("docx", "doc"):
        doc_type = "docx"
        text = extract_text_from_docx(content)
    elif detected_type in ("jpg", "png", "gif", "webp") or file_type_hint in ("jpg", "jpeg", "png", "gif", "webp"):
        doc_type = "image"
        text = extract_text_from_image_bytes(content)
    elif detected_type == "text" or file_type_hint in ("txt", "md", "markdown"):
        doc_type = "txt" if file_type_hint != "md" else "md"
        text = content.decode("utf-8", errors="replace")
    else:
        raise ValueError(f"Tipo de archivo no soportado: {file_type_hint} (detectado: {detected_type})")
    
    chunks = chunk_text(text)
    return text, chunks, doc_type

def extract_text_from_excel(content: bytes) -> str:
    """Extrae texto de un archivo Excel (.xlsx)."""
    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    all_text = []
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_text = [f"[Hoja: {sheet_name}]"]
        
        for row in sheet.iter_rows(values_only=True):
            # Filtrar celdas vacías
            row_values = [str(cell) for cell in row if cell is not None and str(cell).strip()]
            if row_values:
                sheet_text.append(" | ".join(row_values))
        
        if len(sheet_text) > 1:  # Si tiene contenido además del título
            all_text.append("\n".join(sheet_text))
    
    if not all_text:
        raise ValueError("El archivo Excel está vacío o no tiene datos legibles.")
    
    return "\n\n".join(all_text)